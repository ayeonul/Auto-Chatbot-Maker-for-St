import asyncio
import uvicorn
from fastapi import FastAPI, HTTPException
from pydantic import BaseModel
from fastapi.middleware.cors import CORSMiddleware
from typing import Optional

from engine.search import Search
from engine.encoder import Encoder
from utils import *

from openpyxl import Workbook, load_workbook

import os
from dotenv import load_dotenv

load_dotenv()

OPENAI_KEY = os.getenv("OPENAI_KEY")
MIN_INFERENCE_BOUND = float(os.getenv("MIN_INFERENCE_BOUND"))
MAX_INFERENCE_BOUND = float(os.getenv("MAX_INFERENCE_BOUND"))
PORT = int(os.getenv("PORT"))
MULTI_INTENTS_TEXT = os.getenv("MULTI_INTENTS_TEXT")
FALLBACK_TEXT = os.getenv("FALLBACK_TEXT")
ERROR_TEXT = os.getenv("ERROR_TEXT")

app = FastAPI()
origins = [
    "http://localhost:8080",
]
app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


queue = asyncio.Queue()


async def save_to_excel():
    while True:
        user_input = await queue.get()  # queue에서 항목 가져오기

        # 파일이 이미 존재하는 경우
        if os.path.exists("./fallback_log/user_inputs.xlsx"):
            wb = load_workbook("./fallback_log/user_inputs.xlsx")
            ws = wb.active
        # 파일이 존재하지 않는 경우
        else:
            wb = Workbook()
            ws = wb.active

        ws.append([user_input])  # 사용자 입력 추가
        wb.save("./fallback_log/user_inputs.xlsx")


encoder = Encoder()
chatbot_data = ChatbotData(encoder)
engine = Search(chatbot_data.example_embs[["emb"]], chatbot_data.example_embs[["code"]])


chatbot_engine = ChatbotEngine(encoder, engine, chatbot_data.chatbot_contents)


class ChatbotPayload(BaseModel):
    user_input: str


@app.post("/api/chat-engine")
async def chat(payload: ChatbotPayload):
    res = await chatbot_engine.chat(payload.user_input, MAX_INFERENCE_BOUND, MIN_INFERENCE_BOUND, FALLBACK_TEXT, MULTI_INTENTS_TEXT, ERROR_TEXT, queue)
    return res


# func_desc = [
#     {
#         "name": "chatbot",
#         "description": "function that determines the intent of the conversation by finding the embedding similarity with pre-registered example sentences and responds.",
#         "parameters": {
#             "type": "object",
#             "properties": {
#                 "user_input": {
#                     "type": "string",
#                     "description": "texts that condenses the user's utterance intent into a short word or sentence.",
#                 }
#             },
#             "required": ["user_input"],
#         },
#     },
# ]


# def pass_def(user_input):
#     pass


# functions = {"chatbot": pass_def}

# gpt_engine = GPTIntentRecognizer(
#     api_key=OPENAI_KEY,
#     func_desc=func_desc,
#     functions=functions,
# )


# class GPTChatbotPayload(BaseModel):
#     user_input: str
#     messages: list


# @app.post("/api/gpt-chat-engine")
# async def chat(payload: GPTChatbotPayload):
#     user_emb = encoder([payload.user_input]).mean(axis=0, keepdims=True)
#     search_res_df = engine.search(user_emb, n=5)
#     search_res_df = search_res_df.drop_duplicates(["code"]).astype({"code": str})
#     first_res = search_res_df.iloc[0]
#     # print(first_res['score'])

#     if first_res["score"] >= MAX_INFERENCE_BOUND:
#         res_content = find_chat_contents(first_res["code"])
#         if res_content == None:
#             res_content = {
#                 "code": "error",
#                 "title": "error",
#                 "contents": [
#                     {
#                         "type": "speak",
#                         "image": None,
#                         "text": ERROR_TEXT,
#                         "button": [],
#                     }
#                 ],
#             }
#         return res_content

#     elif (
#         MIN_INFERENCE_BOUND < first_res["score"]
#         and first_res["score"] < MAX_INFERENCE_BOUND
#     ):
#         multi_intent_res = {
#             "code": "multi-intents",
#             "title": "multi-intents",
#             "contents": [
#                 {
#                     "type": "speak",
#                     "image": None,
#                     "text": MULTI_INTENTS_TEXT,
#                     "button": [
#                         {
#                             "text": find_chat_contents(data["code"])["title"],
#                             "type": "call_flow",
#                             "action": data["code"],
#                         }
#                         for _, data in search_res_df.iterrows()
#                         if MIN_INFERENCE_BOUND < data["score"]
#                     ],
#                 }
#             ],
#         }
#         return multi_intent_res

#     else:
#         try:
#             short_user_input = gpt_engine(payload.messages, ensure_func_name="chatbot")

#         except:
#             fallback_res = {
#                 "code": "fallback",
#                 "title": "fallback",
#                 "contents": [
#                     {
#                         "type": "speak",
#                         "image": None,
#                         "text": FALLBACK_TEXT,
#                         "button": [],
#                     }
#                 ],
#             }
#             return fallback_res

#         else:
#             user_emb = encoder([short_user_input["user_input"]]).mean(
#                 axis=0, keepdims=True
#             )
#             search_res_df = engine.search(user_emb, n=5)
#             search_res_df = search_res_df.drop_duplicates(["code"]).astype(
#                 {"code": str}
#             )
#             first_res = search_res_df.iloc[0]
#             # print(search_res_df.iloc[0]['score'])

#             if first_res["score"] >= MAX_INFERENCE_BOUND:
#                 res_content = find_chat_contents(first_res["code"])
#                 if res_content == None:
#                     res_content = {
#                         "code": "error",
#                         "title": "error",
#                         "contents": [
#                             {
#                                 "type": "speak",
#                                 "image": None,
#                                 "text": ERROR_TEXT,
#                                 "button": [],
#                             }
#                         ],
#                     }
#                 return res_content

#             elif (
#                 MIN_INFERENCE_BOUND < first_res["score"]
#                 and first_res["score"] < MAX_INFERENCE_BOUND
#             ):
#                 multi_intent_res = {
#                     "code": "multi-intents",
#                     "title": "multi-intents",
#                     "contents": [
#                         {
#                             "type": "speak",
#                             "image": None,
#                             "text": MULTI_INTENTS_TEXT,
#                             "button": [
#                                 {
#                                     "text": find_chat_contents(data["code"]).get(
#                                         "title"
#                                     ),
#                                     "type": "call_flow",
#                                     "action": data["code"],
#                                 }
#                                 for _, data in search_res_df.iterrows()
#                             ],
#                         }
#                     ],
#                 }
#                 return multi_intent_res

#             else:
#                 await queue.put(payload.user_input)
#                 fallback_res = {
#                     "code": "fallback",
#                     "title": "fallback",
#                     "contents": [
#                         {
#                             "type": "speak",
#                             "image": None,
#                             "text": FALLBACK_TEXT,
#                             "button": [],
#                         }
#                     ],
#                 }
#                 return fallback_res


class CallFlowPayload(BaseModel):
    flow_code: str


@app.post("/api/call-flow")
async def call_flow(payload: CallFlowPayload):
    res = chatbot_engine.find_chat_contents(payload.flow_code)
    return res


@app.on_event("startup")
async def startup_event():
    asyncio.create_task(save_to_excel())


async def main():
    config = uvicorn.Config(
        "main:app", host="0.0.0.0", port=PORT, log_level="info", reload=True
    )
    MAIN = uvicorn.Server(config)
    await MAIN.serve()


if __name__ == "__main__":
    asyncio.run(main())
